#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
semicon_tracker.py
==================
반도체 수출/가격/장비 사이클 트래커.

[자동/LIVE] 관세청(TRASS) getNitemtradeList 한 키로 아래를 모두 산출.
  - 메모리(854232) 국가별·제품(10자리)별·HBM프록시(8542323000) 뷰 + ASP/MoM/YoY.
  - MOTIE ICT 동향 재현: 반도체(8541+8542)·디스플레이(8524)·휴대폰(851713) 월별 수출·YoY.
  - DRAM/NAND ASP 프록시(USD/kg) 월별 시계열.  (TrendForce 대체 프록시)

[CSV/누적] 무료 API가 없는 유료·보도자료 소스. 파일에 여러 행을 붙이면 누적·정렬·중복제거.
  - data/trendforce_spot.csv      : DRAM/NAND Spot (TrendForce, 유료 → 직접 입력)
  - data/trendforce_contract.csv  : DRAM/NAND Contract (TrendForce, 유료)
  - data/semi_billings.csv        : SEMI 장비 Billings (보도자료 → 직접 입력)
  - data/trass_speed.csv          : 관세청 10/20일 수출 잠정치 (보도자료)

출력: 멀티시트 .xlsx.  VS Code 에서 아래 상수만 수정 후 Run.
"""

from __future__ import annotations

import argparse
import logging
import sys
import time
import os
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

import pandas as pd
import requests

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s | %(levelname)-7s | %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("semicon")

# ===========================================================================
# ▼▼▼ 여기만 수정 (VS Code 에서 그냥 Run) ▼▼▼
SERVICE_KEY = "9b4b6ada2b7f86c80fc28891bbcea19a7cdedbb5c12be630bd4527ffaff2f918"
ENCODED_KEY = False     # 키에 %2B %2F 등 인코딩 문자가 있으면 True
OUT_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "semicon_tracker.xlsx")
DEMO        = False     # True 면 키 없이 합성데이터로 출력 검증

# --- 실행 모드 ---
# "full"    : MONTHS 개월치를 새로 가져와 캐시에 덮어쓴다.  (최초 1회)
# "update"  : 최근 2개월만 가져와 캐시에 추가한다.          (매월 실행)
MODE   = "update"
MONTHS = 2     # "full" 모드에서만 사용 (2019-01 ~ 현재 ≈ 90개월)
# ▲▲▲ 여기까지 ▲▲▲
# ===========================================================================

TRASS_URL = "https://apis.data.go.kr/1220000/nitemtrade/getNitemtradeList"

HS_CODES: Dict[str, str] = {"854232": "Memory (메모리)"}

PRODUCT_MAP: Dict[str, str] = {
    "8542321010": "DRAM", "8542321020": "SRAM", "8542321030": "NAND Flash",
    "8542321090": "Other Memory", "8542322000": "Hybrid IC",
    "8542323000": "HBM/Advanced (복합구조칩)", "8542324000": "MCOs (복합부품)",
}
HBM_PROXY_CODE = "8542323000"
KEY_COUNTRIES = ["TW", "US", "CN", "HK", "VN", "SG", "JP"]

# MOTIE ICT 재현용 HS 묶음 (수출액 = 각 코드 총계 합산).
ICT_HS: Dict[str, List[str]] = {
    "반도체": ["8541", "8542"],
    "디스플레이": ["8524"],     # HS2022 평판디스플레이모듈
    "휴대폰": ["851713"],       # 스마트폰 포함 셀룰러폰
}

# 반도체 제조장비 수입(HS 8486) = 국내 팹 capex 프록시. (SEMI Billings 자동 대체)
# 한국이 ASML(네덜란드)·TEL(일본)·AMAT/LAM(미국) 장비를 수입 → capex 사이클 선행.
EQUIP_HS: Dict[str, List[str]] = {
    "반도체장비(전공정)": ["848620"],   # 반도체 디바이스/IC 제조장비 = 핵심 capex
    "장비부품·소모성": ["848690"],       # 부품·소모품 = recurring 수요
}
EQUIP_COUNTRY_HS = "848620"          # 장비 향국(공급망) 분해용

# FRED 무료 시계열 (API 키 불필요, CSV 엔드포인트). DRAM/NAND 가격 외부 교차검증.
FRED_SERIES: Dict[str, str] = {
    "미국 반도체 PPI": "PCU334413334413",   # Semiconductor & related device mfg PPI
}
FRED_CSV_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={sid}"

DATA_DIR = Path(os.path.join(os.path.dirname(os.path.abspath(__file__)), "data"))


@dataclass
class Settings:
    service_key: str = SERVICE_KEY
    encoded_key: bool = ENCODED_KEY
    mode: str = MODE        # "full" | "update"
    months: int = MONTHS    # full 모드 전용
    out_path: str = OUT_PATH
    demo: bool = DEMO
    request_pause: float = 0.2
    hs_codes: Dict[str, str] = field(default_factory=lambda: dict(HS_CODES))


# --------------------------------------------------------------------------- #
# Date / parse helpers
# --------------------------------------------------------------------------- #

def month_list(n: int) -> List[str]:
    today = date.today()
    y, m, out = today.year, today.month, []
    for _ in range(n):
        out.append(f"{y}{m:02d}")
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return sorted(out)


def yyyymm_label(s: str) -> str:
    return f"{s[:4]}-{s[4:6]}"


def _to_float(s) -> Optional[float]:
    if s is None or s == "":
        return None
    try:
        return float(str(s).replace(",", ""))
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# Collector: 관세청 TRASS (LIVE)
# --------------------------------------------------------------------------- #

class TrassCollector:
    def __init__(self, cfg: Settings):
        self.cfg = cfg
        self.session = requests.Session()
        self._diag_done = False

    def _get_xml(self, yyyymm: str, hs: str) -> Optional[ET.Element]:
        params = {"strtYymm": yyyymm, "endYymm": yyyymm, "hsSgn": hs}
        if self.cfg.encoded_key:
            url = f"{TRASS_URL}?serviceKey={self.cfg.service_key}"
            resp = self.session.get(url, params=params, timeout=30)
        else:
            params["serviceKey"] = self.cfg.service_key
            resp = self.session.get(TRASS_URL, params=params, timeout=30)
        if resp.status_code != 200:
            body = (resp.text or "")[:300].replace("\n", " ")
            log.error("HTTP %s (%s/%s) | %s", resp.status_code, yyyymm, hs, body)
            self._diagnose()
            return None
        try:
            root = ET.fromstring(resp.text)
        except ET.ParseError:
            log.warning("XML 파싱 실패 (%s/%s)", yyyymm, hs)
            return None
        rc = root.findtext(".//header/resultCode")
        if rc is not None and rc != "00":
            log.error("결과코드 %s (%s/%s): %s", rc, yyyymm, hs, root.findtext(".//header/resultMsg"))
            self._diagnose()
            return None
        return root

    def _diagnose(self):
        if self._diag_done:
            return
        self._diag_done = True
        log.error("-" * 60)
        log.error("진단: 활용신청 승인 여부 / 승인 직후 키 전파 지연(최대 1시간) /")
        log.error("      getNitemtradeList 승인 / https 사용 여부를 확인하세요.")
        log.error("-" * 60)

    def request_items(self, yyyymm: str, hs: str) -> List[dict]:
        """메모리 상세: 국가 × 10자리 행 (총계 제외)."""
        root = self._get_xml(yyyymm, hs)
        if root is None:
            return []
        rows = []
        for item in root.findall(".//item"):
            def g(t): return (item.findtext(t) or "").strip()
            cc, year = g("statCd"), g("year")
            if year == "총계" or cc in ("-", ""):
                continue
            hs10 = g("hsCd")
            rows.append({
                "yyyymm": yyyymm, "month": yyyymm_label(yyyymm), "hs6": hs,
                "hs10": hs10, "category": PRODUCT_MAP.get(hs10, f"기타({hs10})"),
                "country_code": cc, "country_name": g("statCdCntnKor1"),
                "product_kr": g("statKor"),
                "exp_usd": _to_float(g("expDlr")), "exp_wgt": _to_float(g("expWgt")),
                "imp_usd": _to_float(g("impDlr")), "imp_wgt": _to_float(g("impWgt")),
            })
        return rows

    def fetch_total(self, yyyymm: str, hs: str) -> Optional[Tuple[float, float]]:
        """해당 HS 의 월 총계(전세계) 수출액·중량. 총계행 우선, 없으면 합산."""
        root = self._get_xml(yyyymm, hs)
        if root is None:
            return None
        total_usd = total_wgt = None
        sum_usd = sum_wgt = 0.0
        for item in root.findall(".//item"):
            def g(t): return (item.findtext(t) or "").strip()
            if g("year") == "총계" or g("statCd") == "-":
                total_usd = _to_float(g("expDlr"))
                total_wgt = _to_float(g("expWgt"))
            else:
                sum_usd += _to_float(g("expDlr")) or 0
                sum_wgt += _to_float(g("expWgt")) or 0
        if total_usd is not None:
            return total_usd, (total_wgt or 0)
        return (sum_usd, sum_wgt) if sum_usd else None

    def collect_memory(self) -> pd.DataFrame:
        months = month_list(self.cfg.months)
        log.info("TRASS 메모리: %d개월 호출", len(months))
        records, fails = [], 0
        for hs in self.cfg.hs_codes:
            for ym in months:
                rows = self.request_items(ym, hs)
                if rows:
                    records.extend(rows); fails = 0
                    log.info("  %s: %d행", ym, len(rows))
                else:
                    fails += 1
                    if fails >= 3:
                        log.error("연속 3회 빈 응답 → 중단."); return pd.DataFrame.from_records(records)
                time.sleep(self.cfg.request_pause)
        return pd.DataFrame.from_records(records)

    def collect_ict(self) -> pd.DataFrame:
        """MOTIE ICT 재현: 품목별 월 수출액(USD) + YoY/MoM."""
        months = month_list(self.cfg.months)
        n_codes = sum(len(v) for v in ICT_HS.values())
        log.info("ICT 재현: %d개월 × %d HS = %d 호출", len(months), n_codes, len(months) * n_codes)
        rows = []
        for ym in months:
            for item, codes in ICT_HS.items():
                tot, got = 0.0, False
                for hs in codes:
                    t = self.fetch_total(ym, hs)
                    if t:
                        tot += t[0]; got = True
                    time.sleep(self.cfg.request_pause)
                rows.append({"yyyymm": ym, "month": yyyymm_label(ym), "item": item,
                             "exp_usd": tot if got else None})
        df = pd.DataFrame.from_records(rows)
        if df.empty:
            return df
        df["exp_musd"] = df["exp_usd"] / 1e6
        df = df.sort_values(["item", "yyyymm"])
        g = df.groupby("item")
        df["mom_pct"] = g["exp_usd"].pct_change() * 100
        df["yoy_pct"] = g["exp_usd"].pct_change(12) * 100
        return df.reset_index(drop=True)

    def fetch_total_imp(self, yyyymm: str, hs: str) -> Optional[float]:
        """해당 HS 의 월 총계(전세계) 수입액(USD). 총계행 우선, 없으면 합산."""
        root = self._get_xml(yyyymm, hs)
        if root is None:
            return None
        total = None
        s = 0.0
        for item in root.findall(".//item"):
            def g(t): return (item.findtext(t) or "").strip()
            if g("year") == "총계" or g("statCd") == "-":
                total = _to_float(g("impDlr"))
            else:
                s += _to_float(g("impDlr")) or 0
        return total if total is not None else (s if s else None)

    def collect_equipment(self) -> pd.DataFrame:
        """반도체 제조장비(8486) 월별 수입액 = capex 프록시 (카테고리별)."""
        months = month_list(self.cfg.months)
        log.info("반도체장비 수입(8486): %d개월", len(months))
        rows = []
        for ym in months:
            for cat, codes in EQUIP_HS.items():
                usd, got = 0.0, False
                for hs in codes:
                    t = self.fetch_total_imp(ym, hs)
                    if t:
                        usd += t; got = True
                    time.sleep(self.cfg.request_pause)
                rows.append({"yyyymm": ym, "month": yyyymm_label(ym),
                             "category": cat, "imp_usd": usd if got else None})
        return pd.DataFrame.from_records(rows)

    def collect_equipment_country(self) -> pd.DataFrame:
        """반도체장비(848620) 수입 향국별 = 장비 공급망(ASML·TEL·AMAT) 추적.
        848620 아래 여러 HS10이 있으므로 (월·국가) 단위로 합산해 반환(캐시 중복키 보존)."""
        months = month_list(self.cfg.months)
        rec = []
        for ym in months:
            for r in self.request_items(ym, EQUIP_COUNTRY_HS):
                rec.append({"yyyymm": ym, "month": yyyymm_label(ym),
                            "country_code": r["country_code"], "country_name": r["country_name"],
                            "imp_usd": r.get("imp_usd")})
            time.sleep(self.cfg.request_pause)
        df = pd.DataFrame.from_records(rec)
        if df.empty:
            return df
        df["imp_usd"] = pd.to_numeric(df["imp_usd"], errors="coerce")
        return (df.groupby(["yyyymm", "month", "country_code", "country_name"],
                           as_index=False)["imp_usd"].sum())


# --------------------------------------------------------------------------- #
# FRED (무료 거시 시계열, API 키 불필요)
# --------------------------------------------------------------------------- #

def fetch_fred(series_id: str) -> pd.DataFrame:
    """FRED CSV 엔드포인트로 월별 시계열 수집 + MoM/YoY."""
    try:
        resp = requests.get(FRED_CSV_URL.format(sid=series_id), timeout=30)
        if resp.status_code != 200:
            log.warning("FRED %s HTTP %s", series_id, resp.status_code)
            return pd.DataFrame()
        from io import StringIO
        d = pd.read_csv(StringIO(resp.text))
    except Exception as e:
        log.warning("FRED 실패 %s: %s", series_id, e)
        return pd.DataFrame()
    if d.shape[1] < 2:
        return pd.DataFrame()
    d = d.rename(columns={d.columns[0]: "date", d.columns[1]: "value"})
    d["value"] = pd.to_numeric(d["value"], errors="coerce")
    d = d.dropna(subset=["value"]).copy()
    d["yyyymm"] = d["date"].astype(str).str.replace("-", "", regex=False).str[:6]
    d["month"] = d["yyyymm"].apply(lambda s: f"{s[:4]}-{s[4:6]}")
    d["mom_pct"] = d["value"].pct_change() * 100
    d["yoy_pct"] = d["value"].pct_change(12) * 100
    return d[["month", "yyyymm", "value", "mom_pct", "yoy_pct"]].reset_index(drop=True)


def collect_fred(demo: bool = False) -> Dict[str, pd.DataFrame]:
    """FRED_SERIES 전부 수집. 실패 시 data/ 캐시로 폴백."""
    out = {}
    for name, sid in FRED_SERIES.items():
        path = DATA_DIR / f"fred_{sid}.csv"
        df = pd.DataFrame() if demo else fetch_fred(sid)
        if not df.empty:
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            df.to_csv(path, index=False, encoding="utf-8-sig")
            log.info("FRED 수집: %s (%d행)", name, len(df))
        elif path.exists():
            df = pd.read_csv(path)
            log.info("FRED 캐시 사용: %s (%d행)", name, len(df))
        out[name] = df
    return out


# --------------------------------------------------------------------------- #
# Views
# --------------------------------------------------------------------------- #

def _asp(df):
    return df.apply(lambda r: r["exp_usd"] / r["exp_wgt"] if r.get("exp_wgt") else None, axis=1)


def view_country_monthly(raw):
    if raw.empty:
        return raw
    g = (raw.groupby(["yyyymm", "month", "country_code", "country_name"], as_index=False)
            [["exp_usd", "exp_wgt"]].sum())
    g["asp_usd_kg"] = _asp(g)
    g["share_pct"] = g["exp_usd"] / g.groupby("yyyymm")["exp_usd"].transform("sum") * 100
    g = g.sort_values(["country_code", "yyyymm"])
    grp = g.groupby("country_code")
    g["mom_pct"] = grp["exp_usd"].pct_change() * 100
    g["yoy_pct"] = grp["exp_usd"].pct_change(12) * 100
    return g.reset_index(drop=True)


def view_product_monthly(raw):
    if raw.empty:
        return raw
    g = (raw.groupby(["yyyymm", "month", "category"], as_index=False)[["exp_usd", "exp_wgt"]].sum())
    g["asp_usd_kg"] = _asp(g)
    g = g.sort_values(["category", "yyyymm"])
    grp = g.groupby("category")
    g["mom_pct"] = grp["exp_usd"].pct_change() * 100
    g["yoy_pct"] = grp["exp_usd"].pct_change(12) * 100
    return g.reset_index(drop=True)


def view_hbm_by_country(raw):
    if raw.empty:
        return raw
    h = raw[raw["hs10"] == HBM_PROXY_CODE].copy()
    g = (h.groupby(["yyyymm", "month", "country_code", "country_name"], as_index=False)
           [["exp_usd", "exp_wgt"]].sum())
    g["asp_usd_kg"] = _asp(g)
    g = g.sort_values(["country_code", "yyyymm"])
    grp = g.groupby("country_code")
    g["mom_pct"] = grp["exp_usd"].pct_change() * 100
    g["yoy_pct"] = grp["exp_usd"].pct_change(12) * 100
    return g.reset_index(drop=True)


def view_dram_nand_asp(product_m):
    """DRAM/NAND/HBM ASP(USD/kg) 가격 프록시 + 지수(최초=100). TrendForce Spot/Contract 대체."""
    if product_m.empty:
        return product_m
    cats = ["DRAM", "NAND Flash", "HBM/Advanced (복합구조칩)"]
    p = product_m[product_m["category"].isin(cats)].copy().sort_values(["category", "yyyymm"])

    def _index(s):
        base = s.dropna()
        return s / base.iloc[0] * 100 if len(base) else s
    p["asp_index"] = p.groupby("category")["asp_usd_kg"].transform(_index)
    return p[["month", "yyyymm", "category", "exp_usd", "exp_wgt", "asp_usd_kg",
              "asp_index", "mom_pct", "yoy_pct"]].sort_values(
              ["yyyymm", "category"], ascending=[False, True]).reset_index(drop=True)


def enrich_equipment(df):
    """반도체장비 수입 카테고리별 MoM/YoY + 3개월 이동평균(럼피 완화)."""
    if df.empty:
        return df
    df = df.copy().sort_values(["category", "yyyymm"])
    df["imp_musd"] = df["imp_usd"] / 1e6
    g = df.groupby("category")
    df["mom_pct"] = g["imp_usd"].pct_change() * 100
    df["yoy_pct"] = g["imp_usd"].pct_change(12) * 100
    df["ma3_musd"] = g["imp_musd"].transform(lambda s: s.rolling(3, min_periods=1).mean())
    return df.reset_index(drop=True)


def view_equipment_country(df):
    """반도체장비(848620) 수입 향국별 월별 + 비중/MoM/YoY."""
    if df.empty:
        return df
    g = (df.groupby(["yyyymm", "month", "country_code", "country_name"], as_index=False)
           ["imp_usd"].sum())
    g["share_pct"] = g["imp_usd"] / g.groupby("yyyymm")["imp_usd"].transform("sum") * 100
    g = g.sort_values(["country_code", "yyyymm"])
    grp = g.groupby("country_code")
    g["mom_pct"] = grp["imp_usd"].pct_change() * 100
    g["yoy_pct"] = grp["imp_usd"].pct_change(12) * 100
    return g.reset_index(drop=True)


# --------------------------------------------------------------------------- #
# CSV 소스 (누적/정렬/중복제거)
# --------------------------------------------------------------------------- #

CSV_SOURCES = {
    "trass_speed": {"file": "trass_speed.csv", "title": "관세청 10/20일 수출 잠정치 (보도자료 입력)",
                    "sort": "date_label",
                    "header": ["date_label", "item", "exp_musd", "yoy_pct", "note"],
                    "example": ["2025-06 1~10일", "반도체", "5800", "12.5", "보도자료"]},
    "trendforce_spot": {"file": "trendforce_spot.csv", "title": "DRAM/NAND Spot (TrendForce 유료 → 직접 입력)",
                        "sort": "date",
                        "header": ["date", "product", "price_usd", "chg_pct"],
                        "example": ["2025-06-20", "DDR5 16Gb (2Gx8) 5600", "3.85", "-0.5"]},
    "trendforce_contract": {"file": "trendforce_contract.csv", "title": "DRAM/NAND Contract (TrendForce 유료 → 직접 입력)",
                            "sort": "yyyymm",
                            "header": ["yyyymm", "product", "price_usd", "mom_pct"],
                            "example": ["2025-06", "DDR5 16GB Server RDIMM", "180.0", "8.0"]},
    "semi_billings": {"file": "semi_billings.csv", "title": "SEMI 장비 Billings (보도자료 입력, 무료 API 없음)",
                      "sort": "yyyymm",
                      "header": ["yyyymm", "region", "billings_musd", "bookings_musd", "book_to_bill"],
                      "example": ["2025-05", "North America", "4200", "4400", "1.05"]},
}


def _clean_and_sort(df: pd.DataFrame, spec) -> pd.DataFrame:
    """예시행 제거 + 중복제거 + 최신순 정렬."""
    if df.empty:
        return df
    ex = [str(x) for x in spec["example"]]
    df = df[~df.astype(str).apply(lambda r: list(r.values) == ex, axis=1)]
    df = df.drop_duplicates()
    sort_col = spec.get("sort", df.columns[0])
    if sort_col in df.columns:
        df = df.sort_values(sort_col, ascending=False, key=lambda s: s.astype(str))
    return df.reset_index(drop=True)


def load_csv_source(key: str, demo: bool) -> pd.DataFrame:
    spec = CSV_SOURCES[key]
    path = DATA_DIR / spec["file"]
    if not path.exists():
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        pd.DataFrame([spec["example"]], columns=spec["header"]).to_csv(
            path, index=False, encoding="utf-8-sig")
        log.info("템플릿 생성: %s (여러 행 붙이면 누적됨)", path)
        return _demo_csv(key) if demo else pd.DataFrame(columns=spec["header"])
    df = pd.read_csv(path)
    df = _clean_and_sort(df, spec)
    log.info("CSV 로드: %s (%d행)", path, len(df))
    return df


# --------------------------------------------------------------------------- #
# Demo
# --------------------------------------------------------------------------- #

def _demo_memory(cfg):
    import random
    random.seed(7)
    months = month_list(cfg.months)
    countries = [("TW", "대만"), ("US", "미국"), ("CN", "중국"), ("HK", "홍콩"),
                 ("VN", "베트남"), ("SG", "싱가포르"), ("JP", "일본")]
    rows = []
    for i, ym in enumerate(months):
        trend = 1 + 0.02 * i
        for cc, cn in countries:
            for hs10, cat in PRODUCT_MAP.items():
                if cat.startswith("HBM"):
                    base = 3e8 if cc in ("TW", "VN") else 1e7
                elif cat == "DRAM":
                    base = 5e7
                elif cat == "NAND Flash":
                    base = 2e7
                else:
                    base = 5e6
                w = random.uniform(0.6, 1.4)
                usd = base * w * trend
                asp = 113000 if cat.startswith("HBM") else random.uniform(2000, 8000)
                rows.append({"yyyymm": ym, "month": yyyymm_label(ym), "hs6": "854232",
                             "hs10": hs10, "category": cat, "country_code": cc, "country_name": cn,
                             "product_kr": cat, "exp_usd": round(usd), "exp_wgt": round(usd / asp),
                             "imp_usd": round(usd * .3), "imp_wgt": round(usd / asp * .4)})
    return pd.DataFrame.from_records(rows)


def _demo_ict(cfg):
    import random
    random.seed(11)
    months = month_list(cfg.months)
    base = {"반도체": 1.3e10, "디스플레이": 1.5e9, "휴대폰": 1.1e9}
    rows = []
    for i, ym in enumerate(months):
        for item, b in base.items():
            rows.append({"yyyymm": ym, "month": yyyymm_label(ym), "item": item,
                         "exp_usd": b * (1 + .015 * i) * random.uniform(.9, 1.1)})
    df = pd.DataFrame(rows)
    df["exp_musd"] = df["exp_usd"] / 1e6
    df = df.sort_values(["item", "yyyymm"])
    g = df.groupby("item")
    df["mom_pct"] = g["exp_usd"].pct_change() * 100
    df["yoy_pct"] = g["exp_usd"].pct_change(12) * 100
    return df.reset_index(drop=True)


def _demo_csv(key):
    import random
    random.seed(hash(key) % 1000)
    h = CSV_SOURCES[key]["header"]
    if key == "trass_speed":
        return pd.DataFrame([[f"{yyyymm_label(ym)} {t}", "반도체", round(random.uniform(4000, 7000)),
                              round(random.uniform(-10, 30), 1), "demo"]
                             for ym in month_list(6) for t in ("1~10일", "1~20일", "월간")], columns=h)
    if key == "trendforce_spot":
        out, base = [], 3.9
        for ym in month_list(3):
            for d in (5, 12, 19, 26):
                base *= 1 + random.uniform(-.03, .02)
                out.append([f"{yyyymm_label(ym)}-{d:02d}", "DDR5 16Gb (2Gx8) 5600",
                            round(base, 3), round(random.uniform(-3, 2), 2)])
        return pd.DataFrame(out, columns=h)
    if key == "trendforce_contract":
        out, p = [], 175.0
        for ym in month_list(12):
            p *= 1 + random.uniform(-.05, .10)
            out.append([yyyymm_label(ym), "DDR5 16GB Server RDIMM", round(p, 1), round(random.uniform(-5, 12), 1)])
        return pd.DataFrame(out, columns=h)
    if key == "semi_billings":
        out = []
        for ym in month_list(12):
            b = random.uniform(3500, 4800); o = b * random.uniform(.9, 1.2)
            out.append([yyyymm_label(ym), "North America", round(b), round(o), round(o / b, 2)])
        return pd.DataFrame(out, columns=h)
    return pd.DataFrame()


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

NAVY, ACCENT = "1F2A44", "2E5A88"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(color=NAVY, bold=True, size=14)
SECTION_FONT = Font(bold=True, color=NAVY, size=11)
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT, NUM0, NUM2 = "+0.0;-0.0", "#,##0", "#,##0.00"


def _autofit(ws, df, start=1):
    for i, col in enumerate(df.columns):
        w = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(60)]) + 2
        ws.column_dimensions[get_column_letter(start + i)].width = min(w, 30)


def _write_df(ws, df, start_row=1, number_cols=None) -> int:
    number_cols = number_cols or {}
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    for i, (_, rec) in enumerate(df.iterrows(), start_row + 1):
        for j, col in enumerate(df.columns, 1):
            v = rec[col]
            if pd.isna(v):
                v = None
            cell = ws.cell(row=i, column=j, value=v); cell.border = BORDER
            if col in number_cols and v is not None:
                cell.number_format = number_cols[col]
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F6F8FB")
    _autofit(ws, df)
    return start_row + len(df)


def _color_scale(ws, col_letter, first, last):
    if last < first:
        return
    ws.conditional_formatting.add(f"{col_letter}{first}:{col_letter}{last}",
        ColorScaleRule(start_type="num", start_value=-30, start_color="F4B7B0",
                       mid_type="num", mid_value=0, mid_color="FFFFFF",
                       end_type="num", end_value=30, end_color="A7D8B0"))


def build_workbook(raw, country_m, product_m, hbm_m, ict, asp_proxy,
                   csv_data, equip, equip_ctry, fred, out_path):
    wb = Workbook()
    ws = wb.active; ws.title = "Dashboard"
    ws["A1"] = "반도체 수출·가격·장비 사이클 트래커"; ws["A1"].font = TITLE_FONT
    ws["A2"] = f"생성 {date.today().isoformat()}  |  수출 USD · 중량 kg · ASP USD/kg · 8542323000=HBM/첨단패키징 프록시"
    ws["A2"].font = Font(color=ACCENT, size=9, italic=True)
    row = 4

    def section(title, tbl, numcols, scale_cols=()):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        end = _write_df(ws, tbl, row + 1, numcols)
        for col in scale_cols:
            _color_scale(ws, get_column_letter(tbl.columns.get_loc(col) + 1), row + 2, end)
        row = end + 2

    if not country_m.empty:
        latest = country_m["yyyymm"].max()
        snap = country_m[country_m["yyyymm"] == latest].copy()
        snap["__k"] = snap["country_code"].apply(lambda c: KEY_COUNTRIES.index(c) if c in KEY_COUNTRIES else 99)
        snap = snap.sort_values(["__k", "exp_usd"], ascending=[True, False]).head(15)
        tbl = snap[["country_name", "exp_usd", "share_pct", "asp_usd_kg", "mom_pct", "yoy_pct"]].rename(
            columns={"country_name": "향지", "exp_usd": "수출액(USD)", "share_pct": "비중(%)",
                     "asp_usd_kg": "ASP(USD/kg)", "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"})
        section(f"[1] 메모리 수출 국가별 스냅샷  {yyyymm_label(latest)}", tbl,
                {"수출액(USD)": NUM0, "비중(%)": NUM2, "ASP(USD/kg)": NUM0, "MoM(%)": PCT, "YoY(%)": PCT},
                ["MoM(%)", "YoY(%)"])

    if not product_m.empty:
        latest = product_m["yyyymm"].max()
        snap = product_m[product_m["yyyymm"] == latest].sort_values("exp_usd", ascending=False)
        tbl = snap[["category", "exp_usd", "asp_usd_kg", "mom_pct", "yoy_pct"]].rename(
            columns={"category": "제품", "exp_usd": "수출액(USD)", "asp_usd_kg": "ASP(USD/kg)",
                     "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"})
        section(f"[2] 제품 믹스  {yyyymm_label(latest)}", tbl,
                {"수출액(USD)": NUM0, "ASP(USD/kg)": NUM0, "MoM(%)": PCT, "YoY(%)": PCT}, ["YoY(%)"])

    if not hbm_m.empty:
        latest = hbm_m["yyyymm"].max()
        snap = hbm_m[hbm_m["yyyymm"] == latest].sort_values("exp_usd", ascending=False).head(12)
        tbl = snap[["country_name", "exp_usd", "asp_usd_kg", "mom_pct", "yoy_pct"]].rename(
            columns={"country_name": "향지", "exp_usd": "수출액(USD)", "asp_usd_kg": "ASP(USD/kg)",
                     "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"})
        section(f"[3] HBM/첨단패키징 프록시 향지별  {yyyymm_label(latest)}", tbl,
                {"수출액(USD)": NUM0, "ASP(USD/kg)": NUM0, "MoM(%)": PCT, "YoY(%)": PCT}, ["YoY(%)"])

    if not ict.empty:
        latest = ict["yyyymm"].max()
        snap = ict[ict["yyyymm"] == latest].sort_values("exp_usd", ascending=False)
        tbl = snap[["item", "exp_musd", "mom_pct", "yoy_pct"]].rename(
            columns={"item": "품목", "exp_musd": "수출액(백만USD)", "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"})
        section(f"[4] ICT 수출 (관세청 재현)  {yyyymm_label(latest)}", tbl,
                {"수출액(백만USD)": NUM0, "MoM(%)": PCT, "YoY(%)": PCT}, ["YoY(%)"])

    if equip is not None and not equip.empty:
        cur_ym = date.today().strftime("%Y%m")
        avail = sorted(equip["yyyymm"].astype(str).unique())
        comp = [m for m in avail if m < cur_ym]          # 당월(미집계) 제외
        latest = comp[-1] if comp else avail[-1]
        snap = equip[equip["yyyymm"] == latest].sort_values("imp_usd", ascending=False)
        tbl = snap[["category", "imp_musd", "ma3_musd", "mom_pct", "yoy_pct"]].rename(
            columns={"category": "장비", "imp_musd": "수입(백만USD)", "ma3_musd": "3M평균",
                     "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"})
        section(f"[5] 반도체장비 수입 = capex 프록시  {yyyymm_label(latest)}", tbl,
                {"수입(백만USD)": NUM0, "3M평균": NUM0, "MoM(%)": PCT, "YoY(%)": PCT}, ["YoY(%)"])

    fred_rows = []
    for nm, fdf in (fred or {}).items():
        if fdf is None or fdf.empty:
            continue
        last = fdf.sort_values("yyyymm").iloc[-1]
        fred_rows.append({"지표": nm, "기준월": last["month"], "값": round(float(last["value"]), 2),
                          "MoM(%)": last.get("mom_pct"), "YoY(%)": last.get("yoy_pct")})
    if fred_rows:
        section("[6] 외부 가격지표 (FRED · 무료 자동)", pd.DataFrame(fred_rows),
                {"값": NUM2, "MoM(%)": PCT, "YoY(%)": PCT}, ["YoY(%)"])

    ws.column_dimensions["A"].width = 26
    ws.freeze_panes = "A4"

    def sheet(name, df, rename, numcols, note=None):
        if df is None or df.empty:
            return
        w = wb.create_sheet(name[:31]); r0 = 1
        if note:
            w.cell(row=1, column=1, value=note).font = SECTION_FONT; r0 = 3
        out = df.rename(columns=rename)[list(rename.values())]
        _write_df(w, out, r0, numcols)
        w.freeze_panes = w.cell(row=r0 + 1, column=1)

    sheet("국가별_월간", country_m,
          {"month": "월", "country_name": "향지", "exp_usd": "수출액(USD)", "share_pct": "비중(%)",
           "asp_usd_kg": "ASP(USD/kg)", "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"},
          {"수출액(USD)": NUM0, "비중(%)": NUM2, "ASP(USD/kg)": NUM0, "MoM(%)": PCT, "YoY(%)": PCT})
    sheet("제품별_월간", product_m,
          {"month": "월", "category": "제품", "exp_usd": "수출액(USD)", "asp_usd_kg": "ASP(USD/kg)",
           "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"},
          {"수출액(USD)": NUM0, "ASP(USD/kg)": NUM0, "MoM(%)": PCT, "YoY(%)": PCT})
    sheet("HBM프록시_향지별", hbm_m,
          {"month": "월", "country_name": "향지", "exp_usd": "수출액(USD)", "asp_usd_kg": "ASP(USD/kg)",
           "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"},
          {"수출액(USD)": NUM0, "ASP(USD/kg)": NUM0, "MoM(%)": PCT, "YoY(%)": PCT})
    sheet("MOTIE_ICT_월간", ict,
          {"month": "월", "item": "품목", "exp_musd": "수출액(백만USD)", "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"},
          {"수출액(백만USD)": NUM0, "MoM(%)": PCT, "YoY(%)": PCT},
          note="MOTIE ICT 동향 재현 (관세청 8541+8542 / 8524 / 851713 자동 산출)")
    sheet("DRAM_NAND_HBM_ASP_관세청", asp_proxy,
          {"month": "월", "category": "제품", "asp_usd_kg": "ASP(USD/kg)", "asp_index": "ASP지수(최초=100)",
           "exp_usd": "수출액(USD)", "exp_wgt": "중량(kg)", "mom_pct": "ASP MoM(%)", "yoy_pct": "ASP YoY(%)"},
          {"ASP(USD/kg)": NUM0, "ASP지수(최초=100)": NUM2, "수출액(USD)": NUM0, "중량(kg)": NUM0,
           "ASP MoM(%)": PCT, "ASP YoY(%)": PCT},
          note="DRAM/NAND/HBM ASP 가격 프록시 (관세청 USD/kg + 지수, TrendForce Spot/Contract 대체)")
    if not raw.empty:
        sheet("원본_10자리상세", raw,
              {"month": "월", "country_name": "향지", "category": "제품", "hs10": "HS10",
               "product_kr": "품목", "exp_usd": "수출액(USD)", "exp_wgt": "중량(kg)", "imp_usd": "수입액(USD)"},
              {"수출액(USD)": NUM0, "중량(kg)": NUM0, "수입액(USD)": NUM0})

    # --- 자동 대체 소스 (Spot/Contract/Billings 무료 자동화) ---
    sheet("장비수입_월간_capex", equip,
          {"month": "월", "category": "장비", "imp_musd": "수입(백만USD)", "ma3_musd": "3M평균(백만USD)",
           "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"},
          {"수입(백만USD)": NUM0, "3M평균(백만USD)": NUM0, "MoM(%)": PCT, "YoY(%)": PCT},
          note="반도체장비 수입(HS 8486) = 국내 팹 capex 프록시 (SEMI Billings 자동 대체)")
    sheet("장비수입_향국", equip_ctry,
          {"month": "월", "country_name": "수입국", "imp_usd": "수입액(USD)", "share_pct": "비중(%)",
           "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"},
          {"수입액(USD)": NUM0, "비중(%)": NUM2, "MoM(%)": PCT, "YoY(%)": PCT},
          note="반도체장비(848620) 수입 향국 = 공급망 (네덜란드 ASML·일본 TEL·미국 AMAT/LAM)")
    for _nm, _fdf in (fred or {}).items():
        sheet(("FRED_" + _nm.replace(" ", "_"))[:31], _fdf,
              {"month": "월", "value": "지수/값", "mom_pct": "MoM(%)", "yoy_pct": "YoY(%)"},
              {"지수/값": NUM2, "MoM(%)": PCT, "YoY(%)": PCT},
              note=f"{_nm} (FRED 무료 자동수집, API키 불필요)")

    csv_sheet = {"trass_speed": "관세청_10_20일_잠정", "trendforce_spot": "DRAM_NAND_Spot",
                 "trendforce_contract": "DRAM_NAND_Contract", "semi_billings": "SEMI_Billings"}
    for k, df in csv_data.items():
        w = wb.create_sheet(csv_sheet[k][:31])
        w.cell(row=1, column=1, value=CSV_SOURCES[k]["title"]).font = SECTION_FONT
        if df is not None and not df.empty:
            _write_df(w, df, 3)
        else:
            w.cell(row=3, column=1, value=f"data/{CSV_SOURCES[k]['file']} 에 행을 추가하면 누적됩니다.")

    wb.save(out_path)
    log.info("저장 완료: %s", out_path)


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #

# --------------------------------------------------------------------------- #
# Cache (중간 저장소 CSV) — API 응답을 data/ 에 누적 보관
# --------------------------------------------------------------------------- #

CACHE_MEMORY     = DATA_DIR / "cache_memory.csv"
CACHE_ICT        = DATA_DIR / "cache_ict.csv"
CACHE_EQUIP      = DATA_DIR / "cache_equip.csv"
CACHE_EQUIP_CTRY = DATA_DIR / "cache_equip_country.csv"


def _save_cache(df: pd.DataFrame, path: Path):
    """기존 캐시와 머지 후 저장 (yyyymm 기준 중복 제거, 최신순)."""
    if df.empty:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        old = pd.read_csv(path, dtype=str)
        df = pd.concat([old, df.astype(str)], ignore_index=True)
    else:
        df = df.astype(str)
    key_cols = [c for c in ["yyyymm", "hs10", "country_code", "item", "category"] if c in df.columns]
    if key_cols:
        df = df.drop_duplicates(subset=key_cols, keep="last")
    if "yyyymm" in df.columns:
        df = df.sort_values("yyyymm", ascending=False, key=lambda s: s.astype(str))
    df.to_csv(path, index=False, encoding="utf-8-sig")
    log.info("캐시 저장: %s (%d행)", path, len(df))


def _load_cache(path: Path) -> pd.DataFrame:
    if path.exists():
        df = pd.read_csv(path, dtype=str)   # 모든 컬럼 str → yyyymm 숫자 파싱 방지
        # 수치 컬럼은 float 으로 복원
        for col in ["exp_usd", "exp_wgt", "imp_usd", "imp_wgt",
                    "exp_musd", "mom_pct", "yoy_pct"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        log.info("캐시 로드: %s (%d행)", path, len(df))
        return df
    return pd.DataFrame()


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #

def run(cfg: Settings):
    if cfg.demo:
        log.info("=== DEMO MODE ===")
        raw = _demo_memory(cfg)
        ict = _demo_ict(cfg)
        csv_data = {k: _demo_csv(k) for k in CSV_SOURCES}
        equip = pd.DataFrame()
        equip_ctry = pd.DataFrame()
        fred = {}

    else:
        if not cfg.service_key or cfg.service_key.startswith("여기에"):
            log.error("파일 상단 SERVICE_KEY 에 발급키를 넣으세요. (또는 DEMO=True)")
            sys.exit(1)

        tc = TrassCollector(cfg)

        if cfg.mode == "update":
            # ── UPDATE: 최근 2개월만 호출 → 캐시에 추가 ──────────────────
            log.info("=== UPDATE MODE: 최근 2개월 추가 ===")
            cfg_2 = Settings(**{**cfg.__dict__, "months": 2})
            tc2 = TrassCollector(cfg_2)
            new_raw = tc2.collect_memory()
            new_ict = tc2.collect_ict() if not new_raw.empty else pd.DataFrame()
            _save_cache(new_raw, CACHE_MEMORY)
            _save_cache(new_ict, CACHE_ICT)
            _save_cache(tc2.collect_equipment(), CACHE_EQUIP)
            _save_cache(tc2.collect_equipment_country(), CACHE_EQUIP_CTRY)

        else:
            # ── FULL: MONTHS 개월 전체 호출 → 캐시 덮어쓰기 ─────────────
            log.info("=== FULL MODE: %d개월 전체 수집 ===", cfg.months)
            new_raw = tc.collect_memory()
            new_ict = tc.collect_ict() if not new_raw.empty else pd.DataFrame()
            # full 모드는 캐시를 새로 씀 (기존 캐시 무시)
            if not new_raw.empty:
                CACHE_MEMORY.parent.mkdir(parents=True, exist_ok=True)
                new_raw.to_csv(CACHE_MEMORY, index=False, encoding="utf-8-sig")
                log.info("캐시 새로 저장: %s (%d행)", CACHE_MEMORY, len(new_raw))
            if not new_ict.empty:
                CACHE_ICT.parent.mkdir(parents=True, exist_ok=True)
                new_ict.to_csv(CACHE_ICT, index=False, encoding="utf-8-sig")
                log.info("캐시 새로 저장: %s (%d행)", CACHE_ICT, len(new_ict))
            # 장비 수입(8486): full 도 머지(yyyymm·category·향국 키로 중복제거)
            _save_cache(tc.collect_equipment(), CACHE_EQUIP)
            _save_cache(tc.collect_equipment_country(), CACHE_EQUIP_CTRY)

        # 캐시 전체 로드 → 엑셀 생성 (demo 는 위에서 이미 raw/ict 확보)
        raw = _load_cache(CACHE_MEMORY)
        ict = _load_cache(CACHE_ICT)
        csv_data = {k: load_csv_source(k, cfg.demo) for k in CSV_SOURCES}
        equip = enrich_equipment(_load_cache(CACHE_EQUIP))
        equip_ctry = view_equipment_country(_load_cache(CACHE_EQUIP_CTRY))
        fred = collect_fred(cfg.demo)

    country_m = view_country_monthly(raw)
    product_m = view_product_monthly(raw)
    hbm_m = view_hbm_by_country(raw)
    asp_proxy = view_dram_nand_asp(product_m)
    build_workbook(raw, country_m, product_m, hbm_m, ict, asp_proxy,
                   csv_data, equip, equip_ctry, fred, cfg.out_path)

    # 대시보드(industry-dashboard/semiconductor.html) 데이터 JS 자동 생성.
    # 캐시 갱신 + 엑셀 + 대시보드 JS 가 한 번의 Run 으로 끝난다. 실패해도 트래커는 정상 종료.
    try:
        import build_dashboard_data
        build_dashboard_data.build()
    except Exception as e:
        log.warning("대시보드 JS 생성 건너뜀: %s", e)


def parse_args(argv=None) -> Settings:
    d = Settings()
    p = argparse.ArgumentParser()
    p.add_argument("--key", default=d.service_key)
    p.add_argument("--encoded-key", action="store_true", default=d.encoded_key)
    p.add_argument("--mode", default=d.mode, choices=["full", "update"])
    p.add_argument("--months", type=int, default=d.months)
    p.add_argument("--out", default=d.out_path)
    p.add_argument("--demo", action="store_true", default=d.demo)
    a = p.parse_args(argv)
    return Settings(service_key=a.key, encoded_key=a.encoded_key, mode=a.mode,
                    months=a.months, out_path=a.out, demo=a.demo)


if __name__ == "__main__":
    run(parse_args())
