#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sector_tracker.py
=================
국내 상장 섹터별 주요 수출품 트렌드 트래커. (semicon_tracker.py 구조 일반화)

관세청(TRASS) getNitemtradeList 한 키로 아래 8개 섹터를 자동 수집한다.
  K뷰티화장품 · 이차전지 · 자동차 · 철강 · 석유화학 · 조선 · 의료기기 · K푸드

각 섹터마다 두 축으로 산출한다.
  1) 제품믹스(서브카테고리) 월별 수출액·중량·ASP·MoM·YoY  (fetch_total 기반)
  2) 향지(국가)별 월별 수출액·비중·MoM·YoY                 (request_items 합산)

캐시(full/update)와 포맷 엑셀 출력은 반도체 트래커와 동일한 패턴.
HS 코드는 섹터별 SECTORS 딕셔너리에서 관리하며, 일부 소재 코드는 근사치이므로
관세청 'HS CODE 내비게이션'으로 검증 후 조정 권장.

사용: VS Code 에서 아래 상수만 수정 후 Run.
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

import pandas as pd
import requests

# 섹터 정의(HS 코드 매핑)는 sectors_config.py 로 분리 — 엔진과 '리서치 자산' 분리.
# 어디서 Run 해도 임포트되도록 스크립트 폴더를 경로에 추가.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sectors_config import SECTORS

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s | %(levelname)-7s | %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("sector")

# ===========================================================================
# ▼▼▼ 여기만 수정 (VS Code 에서 그냥 Run) ▼▼▼
SERVICE_KEY = "9b4b6ada2b7f86c80fc28891bbcea19a7cdedbb5c12be630bd4527ffaff2f918"
ENCODED_KEY = False
OUT_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output", "sector_tracker.xlsx")
DEMO        = False     # True 면 키 없이 합성데이터로 출력 검증

MODE   = "full"         # "full"(최초 전체) | "update"(매월 +2개월) | "cache"(API 없이 캐시→마스터)
MONTHS = 90             # full 모드 조회 개월 (2019-01 ~ 현재 ≈ 90)

# 돌릴 섹터만 남기면 호출량 절약. 전부면 아래 그대로.
SECTORS_TO_RUN = ["의료기기"]

# 전체 15개:
# "K뷰티화장품", "이차전지", "자동차", "철강", "석유화학", "조선", "의료기기", "K푸드",
# "방산", "전력기기·변압기", "제약·바이오", "엔터·K콘텐츠", "태양광·신재생", "기계·건설기계", "타이어"

# ▲▲▲ 여기까지 ▲▲▲
# ===========================================================================

TRASS_URL = "https://apis.data.go.kr/1220000/nitemtrade/getNitemtradeList"
DATA_DIR = Path(os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_sector"))


@dataclass
class Settings:
    service_key: str = SERVICE_KEY
    encoded_key: bool = ENCODED_KEY
    mode: str = MODE
    months: int = MONTHS
    out_path: str = OUT_PATH
    demo: bool = DEMO
    request_pause: float = 0.15
    sectors: List[str] = field(default_factory=lambda: list(SECTORS_TO_RUN))


# --------------------------------------------------------------------------- #
# Date / parse helpers
# --------------------------------------------------------------------------- #

def month_list(n: int) -> List[str]:
    today = date.today()
    y, m, out = today.year, today.month, []
    for _ in range(n):
        out.append(f"{y}{m:02d}")
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return sorted(out)


def yyyymm_label(s) -> str:
    s = str(s)
    return f"{s[:4]}-{s[4:6]}"


def _to_float(s) -> Optional[float]:
    if s is None or s == "":
        return None
    try:
        return float(str(s).replace(",", ""))
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# 관세청 TRASS Collector
# --------------------------------------------------------------------------- #

class TrassCollector:
    def __init__(self, cfg: Settings):
        self.cfg = cfg
        self.session = requests.Session()
        self._diag_done = False

    def _get_xml(self, yyyymm: str, hs: str) -> Optional[ET.Element]:
        params = {"strtYymm": yyyymm, "endYymm": yyyymm, "hsSgn": hs}
        if self.cfg.encoded_key:
            url = f"{TRASS_URL}?serviceKey={self.cfg.service_key}"
            resp = self.session.get(url, params=params, timeout=30)
        else:
            params["serviceKey"] = self.cfg.service_key
            resp = self.session.get(TRASS_URL, params=params, timeout=30)
        if resp.status_code != 200:
            log.error("HTTP %s (%s/%s) | %s", resp.status_code, yyyymm, hs,
                      (resp.text or "")[:200].replace("\n", " "))
            self._diagnose()
            return None
        try:
            root = ET.fromstring(resp.text)
        except ET.ParseError:
            return None
        rc = root.findtext(".//header/resultCode")
        if rc is not None and rc != "00":
            log.error("결과코드 %s (%s/%s): %s", rc, yyyymm, hs, root.findtext(".//header/resultMsg"))
            self._diagnose()
            return None
        return root

    def _diagnose(self):
        if self._diag_done:
            return
        self._diag_done = True
        log.error("진단: 활용신청 승인/키 전파 지연(최대 1h)/getNitemtradeList 승인/https 확인")

    def fetch_total(self, yyyymm: str, hs: str) -> Optional[Tuple[float, float]]:
        """월 총계(전세계) 수출액·중량."""
        root = self._get_xml(yyyymm, hs)
        if root is None:
            return None
        total_usd = total_wgt = None
        sum_usd = sum_wgt = 0.0
        for it in root.findall(".//item"):
            def g(t): return (it.findtext(t) or "").strip()
            if g("year") == "총계" or g("statCd") == "-":
                total_usd, total_wgt = _to_float(g("expDlr")), _to_float(g("expWgt"))
            else:
                sum_usd += _to_float(g("expDlr")) or 0
                sum_wgt += _to_float(g("expWgt")) or 0
        if total_usd is not None:
            return total_usd, (total_wgt or 0)
        return (sum_usd, sum_wgt) if sum_usd else None

    def fetch_by_country(self, yyyymm: str, hs: str) -> List[dict]:
        """국가별 행 (총계 제외)."""
        root = self._get_xml(yyyymm, hs)
        if root is None:
            return []
        rows = []
        for it in root.findall(".//item"):
            def g(t): return (it.findtext(t) or "").strip()
            cc, year = g("statCd"), g("year")
            if year == "총계" or cc in ("-", ""):
                continue
            rows.append({"yyyymm": yyyymm, "country_code": cc, "country_name": g("statCdCntnKor1"),
                         "hs": hs, "exp_usd": _to_float(g("expDlr")), "exp_wgt": _to_float(g("expWgt"))})
        return rows


# --------------------------------------------------------------------------- #
# 섹터 수집
# --------------------------------------------------------------------------- #

def collect_groups(tc: TrassCollector, sector: str, months: List[str]) -> pd.DataFrame:
    """제품믹스: 서브카테고리 × 월 수출액·중량."""
    cfg = SECTORS[sector]
    rows = []
    for ym in months:
        for label, codes in cfg["groups"].items():
            usd = wgt = 0.0
            got = False
            for hs in codes:
                t = tc.fetch_total(ym, hs)
                if t:
                    usd += t[0]; wgt += t[1]; got = True
                time.sleep(tc.cfg.request_pause)
            rows.append({"sector": sector, "yyyymm": ym, "group": label,
                         "exp_usd": usd if got else None, "exp_wgt": wgt if got else None})
    return pd.DataFrame(rows)


def collect_countries(tc: TrassCollector, sector: str, months: List[str]) -> pd.DataFrame:
    """향지별: 대표 HS 코드 합산 × 월."""
    cfg = SECTORS[sector]
    rec = []
    for ym in months:
        for hs in cfg["country_hs"]:
            rec.extend(tc.fetch_by_country(ym, hs))
            time.sleep(tc.cfg.request_pause)
    if not rec:
        return pd.DataFrame()
    df = pd.DataFrame(rec)
    g = (df.groupby(["yyyymm", "country_code", "country_name"], as_index=False)
           [["exp_usd", "exp_wgt"]].sum())
    g["sector"] = sector
    return g


# --------------------------------------------------------------------------- #
# 파생지표 (ASP/MoM/YoY/share)
# --------------------------------------------------------------------------- #

def _add_ytd(df: pd.DataFrame, key: str) -> pd.DataFrame:
    """연초~당월 누적(ytd_usd)과 전년 동기간 누적 대비(ytd_yoy_pct) 추가.
    럼피(선적 편중) 섹터는 MoM가 노이즈라 이 두 지표로 추세를 판단한다."""
    df = df.sort_values([key, "yyyymm"]).copy()
    df["_yr"] = df["yyyymm"].astype(str).str[:4]
    df["_mo"] = df["yyyymm"].astype(str).str[4:6]
    df["ytd_usd"] = df.groupby([key, "_yr"])["exp_usd"].cumsum()
    # 같은 (key, 월)끼리 연도순 정렬 → pct_change = 전년 동기간 누적 대비
    df = df.sort_values([key, "_mo", "_yr"])
    df["ytd_yoy_pct"] = df.groupby([key, "_mo"])["ytd_usd"].pct_change() * 100
    return df.drop(columns=["_yr", "_mo"])


def enrich_groups(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df["month"] = df["yyyymm"].apply(yyyymm_label)
    df["asp_usd_kg"] = df.apply(lambda r: r["exp_usd"] / r["exp_wgt"]
                                if r.get("exp_wgt") else None, axis=1)
    df = df.sort_values(["group", "yyyymm"])
    gp = df.groupby("group")
    df["mom_pct"] = gp["exp_usd"].pct_change() * 100
    df["yoy_pct"] = gp["exp_usd"].pct_change(12) * 100
    df = _add_ytd(df, "group")
    return df.sort_values(["group", "yyyymm"]).reset_index(drop=True)


def enrich_countries(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df["month"] = df["yyyymm"].apply(yyyymm_label)
    df["share_pct"] = df["exp_usd"] / df.groupby("yyyymm")["exp_usd"].transform("sum") * 100
    df["asp_usd_kg"] = df.apply(lambda r: r["exp_usd"] / r["exp_wgt"]
                                if r.get("exp_wgt") else None, axis=1)
    df = df.sort_values(["country_code", "yyyymm"])
    gp = df.groupby("country_code")
    df["mom_pct"] = gp["exp_usd"].pct_change() * 100
    df["yoy_pct"] = gp["exp_usd"].pct_change(12) * 100
    df = _add_ytd(df, "country_code")
    return df.sort_values(["country_code", "yyyymm"]).reset_index(drop=True)


# --------------------------------------------------------------------------- #
# Cache
# --------------------------------------------------------------------------- #

def _cache_path(sector: str, kind: str) -> Path:
    return DATA_DIR / f"cache_{sector}_{kind}.csv"


def _save_cache(df: pd.DataFrame, path: Path, merge: bool):
    if df.empty:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    df = df.astype(str)
    if merge and path.exists():
        df = pd.concat([pd.read_csv(path, dtype=str), df], ignore_index=True)
    keys = [c for c in ["yyyymm", "group", "country_code"] if c in df.columns]
    if keys:
        df = df.drop_duplicates(subset=keys, keep="last")
    df = df.sort_values("yyyymm", ascending=False, key=lambda s: s.astype(str))
    df.to_csv(path, index=False, encoding="utf-8-sig")
    log.info("캐시 저장: %s (%d행)", path.name, len(df))


def _load_cache(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_csv(path, dtype=str)
    for c in ["exp_usd", "exp_wgt"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


# --------------------------------------------------------------------------- #
# Demo
# --------------------------------------------------------------------------- #

def _demo_groups(sector: str, months: List[str]) -> pd.DataFrame:
    import random
    random.seed(hash(sector) % 9999)
    rows = []
    for i, ym in enumerate(months):
        for label, codes in SECTORS[sector]["groups"].items():
            base = random.uniform(2e7, 6e8)
            usd = base * (1 + 0.012 * i) * random.uniform(0.85, 1.15)
            asp = random.uniform(5, 60)
            rows.append({"sector": sector, "yyyymm": ym, "group": label,
                         "exp_usd": round(usd), "exp_wgt": round(usd / asp)})
    return pd.DataFrame(rows)


def _demo_countries(sector: str, months: List[str]) -> pd.DataFrame:
    import random
    random.seed(hash(sector + "c") % 9999)
    kc = SECTORS[sector]["key_countries"] + ["IN", "BR", "AE"]
    names = {"US": "미국", "CN": "중국", "JP": "일본", "VN": "베트남", "HK": "홍콩",
             "PL": "폴란드", "HU": "헝가리", "DE": "독일", "CA": "캐나다", "MX": "멕시코",
             "AU": "호주", "TW": "대만", "IN": "인도", "TH": "태국", "ID": "인도네시아",
             "TR": "튀르키예", "LR": "라이베리아", "PA": "파나마", "MH": "마셜제도",
             "SG": "싱가포르", "GR": "그리스", "RU": "러시아", "NL": "네덜란드",
             "BR": "브라질", "AE": "아랍에미리트"}
    rows = []
    for i, ym in enumerate(months):
        for cc in kc:
            usd = random.uniform(5e6, 2e8) * (1 + 0.01 * i) * random.uniform(0.8, 1.2)
            asp = random.uniform(5, 60)
            rows.append({"yyyymm": ym, "country_code": cc, "country_name": names.get(cc, cc),
                         "exp_usd": round(usd), "exp_wgt": round(usd / asp), "sector": sector})
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

NAVY, ACCENT = "1F2A44", "2E5A88"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(color=NAVY, bold=True, size=14)
SECTION_FONT = Font(bold=True, color=NAVY, size=11)
SUB_FONT = Font(color=ACCENT, size=9, italic=True)
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT, NUM0, NUM2 = "+0.0;-0.0", "#,##0", "#,##0.00"


def _autofit(ws, df, start=1):
    for i, col in enumerate(df.columns):
        w = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(60)]) + 2
        ws.column_dimensions[get_column_letter(start + i)].width = min(w, 28)


def _write_df(ws, df, start_row=1, number_cols=None) -> int:
    number_cols = number_cols or {}
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    for i, (_, rec) in enumerate(df.iterrows(), start_row + 1):
        for j, col in enumerate(df.columns, 1):
            v = rec[col]
            if pd.isna(v):
                v = None
            cell = ws.cell(row=i, column=j, value=v); cell.border = BORDER
            if col in number_cols and v is not None:
                cell.number_format = number_cols[col]
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F6F8FB")
    _autofit(ws, df)
    return start_row + len(df)


def _color_scale(ws, col_letter, first, last):
    if last < first:
        return
    ws.conditional_formatting.add(f"{col_letter}{first}:{col_letter}{last}",
        ColorScaleRule(start_type="num", start_value=-30, start_color="F4B7B0",
                       mid_type="num", mid_value=0, mid_color="FFFFFF",
                       end_type="num", end_value=30, end_color="A7D8B0"))


def build_workbook(results: Dict[str, dict], out_path: str):
    """results[sector] = {'groups': df, 'countries': df}"""
    wb = Workbook()
    ws = wb.active; ws.title = "Cross_Sector_Dashboard"
    ws["A1"] = "국내 상장 섹터별 수출 트렌드 트래커"; ws["A1"].font = TITLE_FONT
    ws["A2"] = f"생성 {date.today().isoformat()}  |  관세청 TRASS 기반 · 수출 USD · ASP USD/kg · 일부 소재코드 근사"
    ws["A2"].font = SUB_FONT
    row = 4

    # 섹터별 최신월 합계 + YoY/MoM
    # 당월(YYYYMM)은 관세청 집계가 미완성 → 대시보드 기준월에서 제외(직전 완성월 사용).
    # YoY가 -100%로 찍히는 착시 방지. 상세 시트에는 당월 잠정치를 그대로 둔다.
    current_ym = date.today().strftime("%Y%m")
    cross = []
    for sector, d in results.items():
        g = d["groups"]
        if g is None or g.empty:
            continue
        months_avail = sorted(g["yyyymm"].astype(str).unique())
        complete = [m for m in months_avail if m < current_ym]
        latest = complete[-1] if complete else months_avail[-1]
        cur = g[g["yyyymm"] == latest]
        tot = cur["exp_usd"].sum()
        # 섹터 합계 시계열로 YoY
        s_tot = g.groupby("yyyymm")["exp_usd"].sum().sort_index()
        yoy = mom = None
        if latest in s_tot.index:
            idx = list(s_tot.index)
            pos = idx.index(latest)
            if pos >= 1:
                prev = s_tot.iloc[pos - 1]
                mom = (tot / prev - 1) * 100 if prev else None
            if pos >= 12:
                yprev = s_tot.iloc[pos - 12]
                yoy = (tot / yprev - 1) * 100 if yprev else None
        top = cur.sort_values("exp_usd", ascending=False).iloc[0]["group"] if not cur.empty else ""
        # YTD: 연초~기준월 누적(섹터=그룹 합) + 전년 동기간 누적 대비 (럼피 섹터의 핵심 지표)
        ytd = cur["ytd_usd"].sum() if "ytd_usd" in cur.columns else None
        ytd_yoy = None
        if ytd is not None and not pd.isna(ytd):
            ly = f"{int(latest[:4]) - 1}{latest[4:6]}"
            prev_rows = g[g["yyyymm"] == ly]
            if not prev_rows.empty:
                pv = prev_rows["ytd_usd"].sum()
                ytd_yoy = (ytd / pv - 1) * 100 if pv else None
        cross.append({"섹터": sector, "최신월": yyyymm_label(latest),
                      "수출액(USD)": round(tot),
                      "누적(YTD)": round(ytd) if (ytd is not None and not pd.isna(ytd)) else None,
                      "MoM(%)": mom, "YoY(%)": yoy, "YTD YoY(%)": ytd_yoy,
                      "최대품목": top, "대표상장사": SECTORS[sector]["companies"]})
    if cross:
        cdf = pd.DataFrame(cross).sort_values("YoY(%)", ascending=False, na_position="last")
        ws.cell(row=row, column=1, value="[ 섹터 요약 ] 최신월 수출 · 누적(YTD) · 전년대비").font = SECTION_FONT
        end = _write_df(ws, cdf, row + 1,
                        {"수출액(USD)": NUM0, "누적(YTD)": NUM0,
                         "MoM(%)": PCT, "YoY(%)": PCT, "YTD YoY(%)": PCT})
        for cn in ["MoM(%)", "YoY(%)", "YTD YoY(%)"]:
            _color_scale(ws, get_column_letter(cdf.columns.get_loc(cn) + 1), row + 2, end)
    ws.column_dimensions["A"].width = 16
    ws.freeze_panes = "A4"

    # 섹터별 상세 시트
    for sector, d in results.items():
        g, c = d["groups"], d["countries"]

        if g is not None and not g.empty:
            w = wb.create_sheet(f"{sector}_제품"[:31])
            w.cell(row=1, column=1, value=f"{sector} · 제품믹스 월별  ({SECTORS[sector]['companies']})").font = SECTION_FONT
            out = g.rename(columns={"month": "월", "group": "품목", "exp_usd": "수출액(USD)",
                                    "ytd_usd": "누적(YTD)", "exp_wgt": "중량(kg)",
                                    "asp_usd_kg": "ASP(USD/kg)", "mom_pct": "MoM(%)",
                                    "yoy_pct": "YoY(%)", "ytd_yoy_pct": "YTD YoY(%)"})[
                ["월", "품목", "수출액(USD)", "누적(YTD)", "중량(kg)", "ASP(USD/kg)",
                 "MoM(%)", "YoY(%)", "YTD YoY(%)"]]
            end = _write_df(w, out, 3, {"수출액(USD)": NUM0, "누적(YTD)": NUM0, "중량(kg)": NUM0,
                            "ASP(USD/kg)": NUM2, "MoM(%)": PCT, "YoY(%)": PCT, "YTD YoY(%)": PCT})
            _color_scale(w, get_column_letter(out.columns.get_loc("YoY(%)") + 1), 4, end)
            _color_scale(w, get_column_letter(out.columns.get_loc("YTD YoY(%)") + 1), 4, end)
            w.freeze_panes = "A4"

        if c is not None and not c.empty:
            w = wb.create_sheet(f"{sector}_국가"[:31])
            w.cell(row=1, column=1, value=f"{sector} · 향지별 월별 (대표 HS {', '.join(SECTORS[sector]['country_hs'])})").font = SECTION_FONT
            out = c.rename(columns={"month": "월", "country_name": "향지", "exp_usd": "수출액(USD)",
                                    "ytd_usd": "누적(YTD)", "share_pct": "비중(%)",
                                    "asp_usd_kg": "ASP(USD/kg)", "mom_pct": "MoM(%)",
                                    "yoy_pct": "YoY(%)", "ytd_yoy_pct": "YTD YoY(%)"})[
                ["월", "향지", "수출액(USD)", "누적(YTD)", "비중(%)", "ASP(USD/kg)",
                 "MoM(%)", "YoY(%)", "YTD YoY(%)"]]
            end = _write_df(w, out, 3, {"수출액(USD)": NUM0, "누적(YTD)": NUM0, "비중(%)": NUM2,
                            "ASP(USD/kg)": NUM2, "MoM(%)": PCT, "YoY(%)": PCT, "YTD YoY(%)": PCT})
            _color_scale(w, get_column_letter(out.columns.get_loc("YoY(%)") + 1), 4, end)
            _color_scale(w, get_column_letter(out.columns.get_loc("YTD YoY(%)") + 1), 4, end)
            w.freeze_panes = "A4"

    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)
    log.info("저장 완료: %s", out_path)


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #

def discover_cached_sectors() -> List[str]:
    """data_sector/ 의 cache_*_groups.csv 에서 캐시된 섹터명을 추출."""
    out = []
    for p in sorted(DATA_DIR.glob("cache_*_groups.csv")):
        name = p.name[len("cache_"):-len("_groups.csv")]
        if name:
            out.append(name)
    return out


def build_from_cache(cfg: Settings):
    """API 호출 없이 data_sector/ 캐시만 읽어 마스터 엑셀 1개 생성.
    SECTORS_TO_RUN 과 무관하게, 캐시에 존재하는 모든 섹터를 합친다."""
    cached = discover_cached_sectors()
    sectors = [s for s in cached if s in SECTORS]
    orphans = [s for s in cached if s not in SECTORS]
    if orphans:
        log.warning("SECTORS 에 정의 없는 캐시는 건너뜀: %s", ", ".join(orphans))
    if not sectors:
        log.error("data_sector/ 에 사용할 캐시가 없습니다. 먼저 full/update 로 수집하세요.")
        sys.exit(1)
    log.info("=== CACHE MODE: %d개 섹터 (API 호출 없음) → %s ===",
             len(sectors), ", ".join(sectors))
    results: Dict[str, dict] = {}
    for sector in sectors:
        gp, cp = _cache_path(sector, "groups"), _cache_path(sector, "country")
        results[sector] = {
            "groups": enrich_groups(_load_cache(gp)),
            "countries": enrich_countries(_load_cache(cp)),
        }
        log.info("── %s 캐시 로드 ──", sector)
    build_workbook(results, cfg.out_path)


def run(cfg: Settings):
    results: Dict[str, dict] = {}

    if cfg.demo:
        log.info("=== DEMO MODE ===")
        months = month_list(cfg.months)
        for sector in cfg.sectors:
            results[sector] = {
                "groups": enrich_groups(_demo_groups(sector, months)),
                "countries": enrich_countries(_demo_countries(sector, months)),
            }
        build_workbook(results, cfg.out_path)
        return

    if cfg.mode == "cache":
        build_from_cache(cfg)
        return

    if not cfg.service_key or cfg.service_key.startswith("여기에"):
        log.error("파일 상단 SERVICE_KEY 에 발급키를 넣으세요. (또는 DEMO=True)")
        sys.exit(1)

    tc = TrassCollector(cfg)
    is_update = cfg.mode == "update"
    months = month_list(2 if is_update else cfg.months)
    log.info("=== %s MODE: %d개월 × %d섹터 ===",
             cfg.mode.upper(), len(months), len(cfg.sectors))

    for sector in cfg.sectors:
        log.info("── %s 수집 시작 ──", sector)
        g_new = collect_groups(tc, sector, months)
        c_new = collect_countries(tc, sector, months)
        gp, cp = _cache_path(sector, "groups"), _cache_path(sector, "country")
        if is_update:
            _save_cache(g_new, gp, merge=True)
            _save_cache(c_new, cp, merge=True)
        else:
            _save_cache(g_new, gp, merge=False)
            _save_cache(c_new, cp, merge=False)
        results[sector] = {
            "groups": enrich_groups(_load_cache(gp)),
            "countries": enrich_countries(_load_cache(cp)),
        }
        log.info("── %s 완료 ──", sector)

    build_workbook(results, cfg.out_path)


def parse_args(argv=None) -> Settings:
    d = Settings()
    p = argparse.ArgumentParser()
    p.add_argument("--key", default=d.service_key)
    p.add_argument("--encoded-key", action="store_true", default=d.encoded_key)
    p.add_argument("--mode", default=d.mode, choices=["full", "update", "cache"])
    p.add_argument("--months", type=int, default=d.months)
    p.add_argument("--out", default=d.out_path)
    p.add_argument("--demo", action="store_true", default=d.demo)
    p.add_argument("--sectors", nargs="+", default=None,
                   help="수집할 섹터 지정 (미지정 시 SECTORS_TO_RUN). 예: --sectors 방산 타이어")
    a = p.parse_args(argv)
    sectors = a.sectors if a.sectors else list(SECTORS_TO_RUN)
    unknown = [s for s in sectors if s not in SECTORS]
    if unknown:
        p.error(f"알 수 없는 섹터: {unknown}\n가능: {list(SECTORS)}")
    return Settings(service_key=a.key, encoded_key=a.encoded_key, mode=a.mode,
                    months=a.months, out_path=a.out, demo=a.demo, sectors=sectors)


if __name__ == "__main__":
    run(parse_args())
